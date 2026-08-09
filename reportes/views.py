from django.contrib.auth.models import User
from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse

from openpyxl import Workbook

# ReportLab e importaciones para PDFs profesionales
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import cm

from carreras.models import Carrera
from herramientas.models import Herramienta
from inventario.models import Inventario
from permisos.models import Permiso
from sedes.models import Sede
from solicitudes.models import Solicitud
from solicitudes_equipos.models import SolicitudEquipo
from solicitudes_herramientas.models import SolicitudHerramienta


def index(request):
    context = {
        "usuarios": User.objects.count(),
        "carreras": Carrera.objects.count(),
        "sedes": Sede.objects.count(),
        "inventario": Inventario.objects.count(),
        "herramientas": Herramienta.objects.count(),
        "solicitudes_lab": Solicitud.objects.count(),
        "solicitudes_herramientas": SolicitudHerramienta.objects.count(),
        "solicitudes_equipos": SolicitudEquipo.objects.count(),
        "permisos": Permiso.objects.count(),
    }
    return render(
        request,
        "reportes/index.html",
        context
    )


def reporte_inventario(request):
    buscar = request.GET.get("buscar", "")
    inventario = Inventario.objects.all().order_by("-id")

    if buscar:
        inventario = inventario.filter(
            Q(codigo__icontains=buscar) |
            Q(nombre__icontains=buscar) |
            Q(categoria__icontains=buscar) |
            Q(estado__icontains=buscar) |
            Q(sede__nombre__icontains=buscar)
        )

    return render(
        request,
        "reportes/reporte_inventario.html",
        {
            "inventario": inventario,
            "buscar": buscar
        }
    )


def reporte_solicitudes(request):
    buscar = request.GET.get("buscar", "").strip()

    solicitudes = Solicitud.objects.all().order_by("-id")
    herramientas = SolicitudHerramienta.objects.select_related(
        "herramienta"
    ).order_by("-id")
    equipos = SolicitudEquipo.objects.select_related(
        "equipo"
    ).order_by("-id")
    permisos = Permiso.objects.all().order_by("-id")

    if buscar:
        solicitudes = solicitudes.filter(
            Q(estudiante__username__icontains=buscar) |
            Q(estado__icontains=buscar)
        )

        herramientas = herramientas.filter(
            Q(estudiante__icontains=buscar) |
            Q(herramienta__inventario__nombre__icontains=buscar) |
            Q(herramienta__codigo__icontains=buscar) |
            Q(herramienta__marca__icontains=buscar) |
            Q(estado__icontains=buscar)
        )

        equipos = equipos.filter(
            Q(estudiante__icontains=buscar) |
            Q(equipo__nombre__icontains=buscar) |
            Q(estado__icontains=buscar)
        )

        permisos = permisos.filter(
            Q(estudiante__icontains=buscar) |
            Q(estado__icontains=buscar)
        )

    context = {
        "buscar": buscar,
        "solicitudes": solicitudes,
        "herramientas": herramientas,
        "equipos": equipos,
        "permisos": permisos,
    }

    return render(
        request,
        "reportes/reporte_solicitudes.html",
        context
    )


# ==========================================================
# EXPORTAR PDF INDIVIDUALES (BASE GENERADORA Y VISTAS)
# ==========================================================

def _generar_pdf_base(response, titulo_documento, datos_tabla, motivo, observacion=None, datos_adicionales_tabla=None):
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "Titulo",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=19,
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    estilo_subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        alignment=TA_CENTER,
        spaceAfter=12,
    )

    estilo_normal = ParagraphStyle(
        "NormalPersonalizado",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=13,
    )

    estilo_encabezado_tabla = ParagraphStyle(
        "EncabezadoTabla",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=12,
    )

    estilo_valor_tabla = ParagraphStyle(
        "ValorTabla",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
    )

    contenido = []

    # Encabezado Institucional
    contenido.append(Paragraph("INSTITUTO SUPERIOR TECNOLÓGICO TECNOECUATORIANO", estilo_titulo))
    contenido.append(Paragraph("SISTEMA INSTITUCIONAL EVA2", estilo_subtitulo))
    contenido.append(Paragraph(titulo_documento, estilo_titulo))
    contenido.append(Spacer(1, 0.4 * cm))

    # Formatear celdas envolviendo todo en Paragraphs para wrapping correcto
    datos_procesados = []
    for fila in datos_tabla:
        fila_proc = []
        for i, celda in enumerate(fila):
            if i % 2 == 0:
                # Etiquetas (Columna 0 y 2)
                fila_proc.append(Paragraph(str(celda), estilo_encabezado_tabla))
            else:
                # Valores (Columna 1 y 3)
                fila_proc.append(Paragraph(str(celda), estilo_valor_tabla))
        datos_procesados.append(fila_proc)

    # Anchos optimizados para A4 (Ancho total disponible: 18 cm)
    tabla_datos = Table(
        datos_procesados, 
        colWidths=[2.8 * cm, 6.2 * cm, 2.8 * cm, 6.2 * cm]
    )
    
    estilo_tabla = TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E9F2F9")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#E9F2F9")),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#B8C7D1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])
    tabla_datos.setStyle(estilo_tabla)
    contenido.append(tabla_datos)
    contenido.append(Spacer(1, 0.5 * cm))

    # Sección Motivo
    contenido.append(Paragraph("MOTIVO DE LA SOLICITUD", estilo_encabezado_tabla))
    contenido.append(Spacer(1, 0.15 * cm))
    motivo_texto = motivo or "No especificado."
    contenido.append(Paragraph(str(motivo_texto).replace("\n", "<br/>"), estilo_normal))
    contenido.append(Spacer(1, 0.5 * cm))

    # Sección Observación (si aplica)
    if observacion is not None:
        contenido.append(Paragraph("OBSERVACIÓN", estilo_encabezado_tabla))
        contenido.append(Spacer(1, 0.15 * cm))
        obs_texto = observacion or "Ninguna."
        contenido.append(Paragraph(str(obs_texto).replace("\n", "<br/>"), estilo_normal))
        contenido.append(Spacer(1, 0.5 * cm))

    # Sección de Respuesta / Adicional
    if datos_adicionales_tabla:
        datos_resp_proc = []
        for fila in datos_adicionales_tabla:
            datos_resp_proc.append([
                Paragraph(str(fila[0]), estilo_encabezado_tabla), 
                Paragraph(str(fila[1]), estilo_valor_tabla)
            ])
        tabla_resp = Table(datos_resp_proc, colWidths=[4.5 * cm, 13.5 * cm])
        tabla_resp.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E9F2F9")),
            ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#B8C7D1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        contenido.append(tabla_resp)
        contenido.append(Spacer(1, 0.5 * cm))

    contenido.append(Spacer(1, 0.4 * cm))
    contenido.append(
        Paragraph(
            "Documento generado automáticamente por el Sistema Institucional EVA2.",
            ParagraphStyle(
                "Pie",
                parent=estilos["Normal"],
                fontName="Helvetica",
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER,
            )
        )
    )

    doc.build(contenido)
    return response


def exportar_pdf_solicitud(request, id):
    solicitud = get_object_or_404(Solicitud, id=id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Solicitud_Laboratorio_{solicitud.id}.pdf"'

    estudiante = solicitud.estudiante.get_full_name() or solicitud.estudiante.username

    datos = [
        ["N.º de solicitud", solicitud.id, "Estado", solicitud.estado],
        ["Estudiante", estudiante, "Carrera", solicitud.carrera],
        ["Sede", solicitud.sede, "Laboratorio", solicitud.laboratorio],
        ["Docente", solicitud.docente, "Fecha", solicitud.fecha.strftime("%d/%m/%Y")],
        ["Hora de inicio", solicitud.hora_inicio, "Hora de finalización", solicitud.hora_fin],
    ]

    datos_respuesta = None
    if getattr(solicitud, "aprobado_por", None):
        aprobado = solicitud.aprobado_por.get_full_name() or solicitud.aprobado_por.username
        f_resp = solicitud.fecha_respuesta.strftime("%d/%m/%Y %H:%M") if getattr(solicitud, "fecha_respuesta", None) else "No registrada"
        datos_respuesta = [
            ["Respondido por", aprobado],
            ["Fecha de respuesta", f_resp]
        ]

    return _generar_pdf_base(
        response=response,
        titulo_documento="SOLICITUD DE LABORATORIO",
        datos_tabla=datos,
        motivo=solicitud.motivo,
        observacion=solicitud.observacion,
        datos_adicionales_tabla=datos_respuesta
    )


def exportar_pdf_herramienta(request, id):
    solicitud = get_object_or_404(SolicitudHerramienta, id=id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Solicitud_Herramienta_{solicitud.id}.pdf"'

    f_inicio = solicitud.fecha_inicio.strftime("%d/%m/%Y") if solicitud.fecha_inicio else "N/A"
    f_fin = solicitud.fecha_fin.strftime("%d/%m/%Y") if solicitud.fecha_fin else "N/A"

    datos = [
        ["N.º de solicitud", solicitud.id, "Estado", solicitud.estado],
        ["Estudiante", solicitud.estudiante, "Carrera", solicitud.carrera],
        ["Docente/Profesor", solicitud.profesor, "Herramienta", solicitud.herramienta],
        ["Cantidad", solicitud.cantidad, "Fecha Inicio", f_inicio],
        ["Fecha Fin", f_fin, "Fecha Registro", solicitud.fecha_registro.strftime("%d/%m/%Y %H:%M")],
    ]

    return _generar_pdf_base(
        response=response,
        titulo_documento="SOLICITUD DE HERRAMIENTA",
        datos_tabla=datos,
        motivo=solicitud.motivo
    )


def exportar_pdf_equipo(request, id):
    solicitud = get_object_or_404(SolicitudEquipo, id=id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Solicitud_Equipo_{solicitud.id}.pdf"'

    f_inicio = solicitud.fecha_inicio.strftime("%d/%m/%Y") if solicitud.fecha_inicio else "N/A"
    f_fin = solicitud.fecha_fin.strftime("%d/%m/%Y") if solicitud.fecha_fin else "N/A"

    datos = [
        ["N.º de solicitud", solicitud.id, "Estado", solicitud.estado],
        ["Estudiante", solicitud.estudiante, "Carrera", solicitud.carrera],
        ["Docente/Profesor", solicitud.profesor, "Equipo", solicitud.equipo],
        ["Cantidad", solicitud.cantidad, "N.º de Equipo", solicitud.numero_equipo or "N/A"],
        ["N.º de Serie", solicitud.numero_serie or "N/A", "Fecha Inicio", f_inicio],
        ["Fecha Fin", f_fin, "Fecha Registro", solicitud.fecha_registro.strftime("%d/%m/%Y %H:%M")],
    ]

    return _generar_pdf_base(
        response=response,
        titulo_documento="SOLICITUD DE EQUIPO",
        datos_tabla=datos,
        motivo=solicitud.motivo
    )


def exportar_pdf_permiso(request, id):
    solicitud = get_object_or_404(Permiso, id=id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Permiso_{solicitud.id}.pdf"'

    f_inicio = solicitud.fecha_inicio.strftime("%d/%m/%Y") if solicitud.fecha_inicio else "N/A"
    f_fin = solicitud.fecha_fin.strftime("%d/%m/%Y") if solicitud.fecha_fin else "N/A"

    datos = [
        ["N.º de Permiso", solicitud.id, "Estado", solicitud.estado],
        ["Estudiante", solicitud.estudiante, "Carrera", solicitud.carrera],
        ["Docente", solicitud.docente, "Fecha Inicio", f_inicio],
        ["Fecha Fin", f_fin, "Fecha Solicitud", solicitud.fecha_solicitud.strftime("%d/%m/%Y %H:%M")],
    ]

    return _generar_pdf_base(
        response=response,
        titulo_documento="SOLICITUD DE PERMISO",
        datos_tabla=datos,
        motivo=solicitud.motivo
    )


# ==========================================================
# EXPORTAR INVENTARIO Y SOLICITUDES A EXCEL
# ==========================================================

def exportar_inventario_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append([
        "ID",
        "Código",
        "Nombre",
        "Categoría",
        "Carrera",
        "Sede",
        "Cantidad",
        "Estado"
    ])

    inventario = Inventario.objects.all().order_by("id")

    for item in inventario:
        ws.append([
            item.id,
            item.codigo,
            item.nombre,
            item.categoria,
            str(item.carrera),
            str(item.sede),
            item.cantidad,
            item.estado,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="Reporte_Inventario.xlsx"'
    )

    wb.save(response)
    return response


def exportar_solicitudes_excel(request):
    wb = Workbook()

    # LABORATORIOS
    ws = wb.active
    ws.title = "Laboratorios"
    ws.append(["ID", "Estudiante", "Laboratorio", "Fecha", "Estado"])

    for s in Solicitud.objects.all().order_by("id"):
        ws.append([
            s.id,
            str(s.estudiante),
            str(s.laboratorio),
            s.fecha.strftime("%d/%m/%Y"),
            s.estado
        ])

    # HERRAMIENTAS
    ws2 = wb.create_sheet("Herramientas")
    ws2.append(["ID", "Estudiante", "Herramienta", "Cantidad", "Estado"])

    for h in SolicitudHerramienta.objects.all().order_by("id"):
        ws2.append([
            h.id,
            h.estudiante,
            str(h.herramienta),
            h.cantidad,
            h.estado
        ])

    # EQUIPOS
    ws3 = wb.create_sheet("Equipos")
    ws3.append(["ID", "Estudiante", "Equipo", "Cantidad", "Estado"])

    for e in SolicitudEquipo.objects.all().order_by("id"):
        ws3.append([
            e.id,
            e.estudiante,
            str(e.equipo),
            e.cantidad,
            e.estado
        ])

    # PERMISOS
    ws4 = wb.create_sheet("Permisos")
    ws4.append(["ID", "Estudiante", "Fecha Inicio", "Fecha Fin", "Estado"])

    for p in Permiso.objects.all().order_by("id"):
        ws4.append([
            p.id,
            p.estudiante,
            p.fecha_inicio.strftime("%d/%m/%Y"),
            p.fecha_fin.strftime("%d/%m/%Y"),
            p.estado
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="Reporte_Solicitudes.xlsx"'
    )

    wb.save(response)
    return response


# ==========================================================
# EXPORTAR WORD Y GENERALES
# ==========================================================

def exportar_inventario_word(request):
    return HttpResponse("Próximamente disponible.")


def exportar_solicitudes_word(request):
    return HttpResponse("Próximamente disponible.")


def exportar_inventario_pdf(request):
    return HttpResponse("Próximamente disponible.")


def exportar_solicitudes_pdf(request):
    return HttpResponse("Próximamente disponible.")